from flask import Flask, request, session, redirect, url_for, render_template, flash, jsonify, send_file, Response
from ldap3 import Server, Connection, ALL_ATTRIBUTES, SUBTREE
import integration
from notification import get_user_notifications, get_db_connection
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO, TextIOWrapper
import uuid
import asyncio
import json
import os
import unicodedata
from unidecode import unidecode
import re
from datetime import datetime
from unidecode import unidecode
from maps.Maps import realizar_pesquisas
from server.errorLog import get_error_logs
from server.loginLog import get_login_logs
from api.PostApiData import criar_negocio
from api.CrmDealList import verificar_negocio_existente
import locale
import sys
import io
from auth.login import authenticate

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

app = Flask(__name__)
app.secret_key = 'calvo'

def remover_acentos(texto):
    return unidecode(texto)

def get_user_session_file():
    # Gera um caminho para o arquivo de cache com base no UUID do usuário
    if 'user_id' not in session:
        session['user_id'] = str(uuid.uuid4())

    user_id = session['user_id']
    file_path = os.path.join('cache', f"{user_id}.json")

    # Cria o arquivo temporário se não existir
    if not os.path.exists(file_path):
        with open(file_path, 'w') as file:
            json.dump([], file)

    return file_path

def gerar_excel(empresas, resultados_consolidados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    # Títulos personalizados e estilizados
    titulos = [
        'Razão Social', 'Nome Fantasia', 'CNPJ', 'Capital Social', 'Endereço',
        'Telefone 1', 'Telefone 2', 'Telefone 3', 'Telefone 4', 'E-mail', 'Site', 'Redes Sociais'
    ]

    # Identificar e adicionar títulos para cada sócio dinamicamente
    num_socios = max(len(emp.get('socios', [])) for emp in empresas)
    for i in range(1, num_socios + 1):
        titulos.extend([f'Sócio{i} Nome', f'Sócio{i} Qualificação'])

    # Estilo dos títulos
    for col_num, titulo in enumerate(titulos, 1):
        cell = ws.cell(row=1, column=col_num, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adicionar dados de cada empresa
    for row_num, (empresa, resultado) in enumerate(zip(empresas, resultados_consolidados), start=2):
        ws.cell(row=row_num, column=1, value=empresa.get('razao_social'))
        ws.cell(row=row_num, column=2, value=empresa.get('nome_fantasia'))
        ws.cell(row=row_num, column=3, value=format_cnpj(empresa.get('cnpj')))
        ws.cell(row=row_num, column=4, value=formata_real(empresa.get('capital_social')))
        ws.cell(row=row_num, column=5, value=empresa.get('logradouro'))

        # Telefones e e-mail
        ws.cell(row=row_num, column=6, value=resultado.get('phone1'))
        ws.cell(row=row_num, column=7, value=resultado.get('phone2'))
        ws.cell(row=row_num, column=8, value=empresa.get('telefone_1'))
        ws.cell(row=row_num, column=9, value=empresa.get('telefone_2'))
        ws.cell(row=row_num, column=10, value=empresa.get('email'))

        # Site e Redes Sociais
        ws.cell(row=row_num, column=11, value=resultado.get('website'))
        redes_sociais = ', '.join(resultado.get('social_media_profiles', [])) if resultado.get('social_media_profiles') else None
        ws.cell(row=row_num, column=12, value=redes_sociais)

        # Sócios
        for i, socio in enumerate(empresa.get('socios', []), start=1):
            ws.cell(row=row_num, column=12 + 2 * i - 1, value=socio['nome'])
            ws.cell(row=row_num, column=12 + 2 * i, value=socio['qualificacao'])

    # Ajustar largura das colunas usando a função auxiliar
    ajustar_largura_colunas(ws)

    # Salvar e retornar o arquivo em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Função para ajustar a largura das colunas
def ajustar_largura_colunas(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def carregar_blacklist():
    blacklist = {}
    try:
        with open('resource/blacklist.csv', 'r', newline='') as blacklist_file:
            reader = csv.DictReader(blacklist_file)
            for row in reader:
                cnpj = row['cnpj']
                user = row['user']
                datetime_str = row['datetime']
                blacklist[cnpj] = {
                    'user': user,
                    'datetime': datetime_str
                }
    except FileNotFoundError:
        print("O arquivo blacklist.csv não foi encontrado.")
    return blacklist

@app.after_request
def add_cache_control_headers(resp):
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.route('/', methods=['GET', 'POST'])
def login():
    if 'logged_in' in session and session['logged_in']:
        return redirect(url_for('home'))

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if authenticate(username, password):
            return redirect(url_for('home'))
        else:
            flash('Credenciais inválidas. Tente novamente.', 'error')

    return render_template('login.html')

@app.template_filter('format_cnpj')
def format_cnpj(cnpj):
    if cnpj and len(cnpj) == 14:  # Verifica se o CNPJ tem 14 dígitos
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj

@app.template_filter('format_datetime')
def format_datetime(dt_str):
    if dt_str:
        try:
            dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            return dt.strftime("%d/%m/%Y às %H:%M h")
        except ValueError:
            return dt_str
    return ""

@app.template_filter('formata_real')
def formata_real(valor):
    try:
        valor_float = float(valor) if isinstance(valor, str) else valor
        return locale.currency(valor_float, grouping=True, symbol="R$")
    except (TypeError, ValueError):
        return "R$0,00"

@app.template_filter('capitalize')
def capitalize(value):
    if value.isupper() and len(value) == 2:  # Manter UF em maiúsculas
        return value
    else:
        return value.capitalize()

@app.template_filter('format_cep')
def format_cep(value):
    if value and len(value) == 8 and value.isdigit():
        return f"{value[:5]}-{value[5:]}"
    return value

@app.template_filter('format_phone')
def format_phone(value):
    if value and value.isdigit():
        if len(value) == 10:
            ddd = value[:2]
            prefixo = value[2:6]
            sufixo = value[6:]
            if prefixo[0] in '6 7 8 9':
                return f"({ddd}) 9{prefixo}-{sufixo}"
            return f"({ddd}) {prefixo}-{sufixo}"
        elif len(value) == 11:
            ddd = value[:2]
            prefixo = value[2:7]
            sufixo = value[7:]
            return f"({ddd}) {prefixo}-{sufixo}"
    return value

@app.route('/home', methods=['GET'])
def home():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    full_name = session.get('full_name', 'Usuário')
    username = session.get('username', 'Usuário')
    is_admin = session.get('is_admin', False)

    notifications = get_user_notifications(username)
    show_notification = request.args.get('show_notification', 'true') == 'true'

    termo_busca = request.args.get('termo_busca', '')
    estado = request.args.get('estado', '')
    cidade = request.args.get('cidade', '')
    bairro = request.args.get('bairro', '')
    page = int(request.args.get('page', 1))

    estado = remover_acentos(estado)
    cidade = remover_acentos(cidade)
    bairro = remover_acentos(bairro)

    dados_cnpj = []
    tem_mais_paginas = False
    erro = None
    total_results = 0

    # Carregar dados da blacklist
    blacklist = carregar_blacklist()

    if termo_busca:
        dados_cnpj, total_results, erro = integration.obter_dados_cnpj(termo_busca, estado, cidade, bairro, page)
        print(f"Total de resultados encontrados: {total_results}")

        if dados_cnpj:
            for empresa in dados_cnpj:
                cnpj = empresa.get('cnpj')
                detalhes = integration.obter_detalhes_cnpj(cnpj)
                if detalhes:
                    # Atualizar o dicionário empresa com os detalhes, incluindo capital social
                    empresa.update(detalhes)

                # Adicionar informações da blacklist, se existir
                if cnpj in blacklist:
                    empresa['blacklist_info'] = blacklist[cnpj]

            # Ordenar pelo nome da razão social
            dados_cnpj.sort(key=lambda empresa: empresa.get('razao_social', '').lower())

            # Verificar se existe próxima página de resultados
            proxima_pagina = integration.obter_dados_cnpj(termo_busca, estado, cidade, bairro, page + 1)
            tem_mais_paginas = bool(proxima_pagina)
        else:
            if not erro:
                flash('Nenhum dado encontrado para o termo informado.', 'error')

    if erro:
        flash(erro, 'error')

    return render_template(
        'index.html',
        full_name=full_name,
        notifications=notifications,
        show_notification=show_notification,
        dados_cnpj=dados_cnpj,
        termo_busca=termo_busca,
        estado=estado,
        cidade=cidade,
        bairro=bairro,
        page=page,
        tem_mais_paginas=tem_mais_paginas,
        total_results=total_results,
        is_admin=is_admin
    )


@app.route('/admin_dashboard')
def admin_dashboard():
    # Verificar se o usuário está logado e se é administrador
    if not session.get('logged_in') or not session.get('is_admin'):
        flash('Acesso negado. Você não tem permissão para acessar essa página.', 'error')
        return redirect(url_for('home'))

    # Pegar os logs de erro
    error_logs = get_error_logs()

    # Renderizar os logs no template
    return render_template('admin_dashboard.html', logs=error_logs)

@app.route('/error_log')
def error_log():
    # Verificar se o usuário está logado e se é administrador
    if not session.get('logged_in') or not session.get('is_admin'):
        flash('Acesso negado. Você não tem permissão para acessar essa página.', 'error')
        return redirect(url_for('home'))

    # Pegar os logs de erro
    error_logs = get_error_logs()

    # Renderizar os logs no template
    return render_template('error_log.html', logs=error_logs)

@app.route('/login_log')
def login_log():
    # Verificar se o usuário está logado e se é administrador
    if not session.get('logged_in') or not session.get('is_admin'):
        flash('Acesso negado. Você não tem permissão para acessar essa página.', 'error')
        return redirect(url_for('home'))

    # Pegar os logs de login
    login_logs = get_login_logs()

    # Renderizar os logs no template
    return render_template('login_log.html', logs=login_logs)

@app.route('/dismiss_notification', methods=['POST'])
def dismiss_notification():
    if not session.get('logged_in'):
        return jsonify({'error': 'Usuário não autenticado'}), 401

    notification_id = request.form.get('notification_id')
    accepted = request.form.get('accepted') == 'true'

    if notification_id and accepted:
        try:
            # Atualizar o status enviado_bitrix para true
            connection = get_db_connection()
            cursor = connection.cursor()
            cursor.execute("""
                UPDATE "Result"
                SET enviado_bitrix = true
                WHERE id = %s
            """, (notification_id,))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'success': True}), 200
        except Exception as e:
            print(f"Erro ao atualizar notificação: {e}")
            return jsonify({'error': 'Erro ao atualizar notificação'}), 500
    else:
        return jsonify({'error': 'ID de notificação inválido ou ação não especificada'}), 400


@app.route('/criar_negocio', methods=['GET'])
def criar_negocio_para_empresas():
    try:
        file_path = get_user_session_file()

        # Lê os dados do arquivo temporário
        with open(file_path, 'r') as file:
            empresas = json.load(file)

        if not empresas:
            return jsonify({'error': 'Nenhuma empresa disponível para criar negócio.', 'links': [], 'razoes_sociais': []}), 400

        empresas_para_scrap = []
        links_existentes = []  # Lista para armazenar os links dos negócios existentes
        razoes_sociais_existentes = []  # Lista para armazenar as razões sociais

        # Verifica se cada empresa já tem um negócio no Bitrix, se não tiver, adiciona para scrap
        for empresa in empresas:
            cnpj = empresa.get('cnpj', '')
            razao_social = empresa.get('razao_social', '')

            # Verifica se o negócio já existe com o CNPJ
            negocio_existe, lead_link = verificar_negocio_existente(cnpj)

            if negocio_existe:
                # Verifica se o link é válido
                if lead_link and lead_link.startswith("http"):
                    print(f"Negócio já existente. Link: {lead_link}", flush=True)
                    # Armazena o link e a razão social da empresa
                    links_existentes.append(lead_link)
                    razoes_sociais_existentes.append(razao_social)
                else:
                    # Se o link for inválido, loga a informação e continua
                    print(f"Link inválido ignorado para a empresa {razao_social}: {lead_link}", flush=True)
            else:
                empresas_para_scrap.append(empresa)

        if not empresas_para_scrap:
            print("Todos os negócios já existem no Bitrix.", flush=True)

            # Remove o arquivo de cache antes de retornar
            os.remove(file_path)
            session.pop('user_id', None)

            return jsonify({
                'message': 'Todos os negócios já existem no Bitrix.',
                'links': links_existentes,
                'razoes_sociais': razoes_sociais_existentes
            }), 200

        # Faz o scrap de forma assíncrona apenas para as empresas que não possuem negócio
        queries = [f"{empresa.get('nome_fantasia', empresa.get('razao_social', ''))} {empresa.get('municipio', '')}" for empresa in empresas_para_scrap]

        # Usa asyncio.run em vez de criar um novo loop
        scrap_results_google = asyncio.run(integration.scrap.process_queries(queries))
        scrap_results_maps = realizar_pesquisas(queries)
        scrap_results = scrap_results_consolidated(scrap_results_maps, scrap_results_google)

        # Para cada empresa, cria um novo negócio no Bitrix24 e armazena o link
        for i, empresa in enumerate(empresas_para_scrap):
            razao_social = empresa.get('razao_social', '')
            nome_fantasia = empresa.get('nome_fantasia', '')
            cnpj = empresa.get('cnpj', '')
            endereco = empresa.get('endereco', '')
            email = empresa.get('email', '')
            capital_social = empresa.get('capital_social', '')
            socios = empresa.get('socios', '')

            # Pega os telefones e outras informações do scrap_results
            telefone1 = scrap_results[i].get('phone1', '')
            telefone2 = scrap_results[i].get('phone2', '')
            telefone3 = scrap_results[i].get('phone3', '')  # Caso exista no `scrap_results`
            social_media_profiles = scrap_results[i].get('social_media_profiles', '')
            website = scrap_results[i].get('website', '')

            # Cria o negócio e obtém o link, passando o novo parâmetro social_media_profiles e website
            link_negocio = criar_negocio(
                razao_social, nome_fantasia, cnpj, endereco,
                telefone1, telefone2, telefone3, email, capital_social, socios,
                scrap_results[i], social_media_profiles, website
            )

            # Verifica se o link é válido antes de adicionar
            if link_negocio and link_negocio.startswith("http"):
                links_existentes.append(link_negocio)
                razoes_sociais_existentes.append(razao_social)
            else:
                print(f"Link inválido ignorado para a empresa {razao_social}: {link_negocio}", flush=True)


        # Remove o arquivo de cache e limpa a sessão após a execução bem-sucedida
        os.remove(file_path)
        session.pop('user_id', None)

        return jsonify({
            'message': 'Negócios criados com sucesso!',
            'links': links_existentes,
            'razoes_sociais': razoes_sociais_existentes
        }), 200

    except Exception as e:
        # Em caso de erro, o arquivo não será removido, e o erro será retornado
        print(f"Erro ao processar o negócio: {str(e)}", flush=True)
        return jsonify({'error': 'Erro ao processar os negócios.', 'message': str(e)}), 500


@app.route('/salvar_csv', methods=['POST'])
def salvar_csv():
    form_data = request.form.to_dict(flat=False)

    empresa = {
        "razao_social": form_data.get('razao_social', [''])[0],
        "nome_fantasia": form_data.get('nome_fantasia', [''])[0],
        "logradouro": form_data.get('logradouro', [''])[0],
        "telefone_1": form_data.get('telefone_1', [''])[0],
        "telefone_2": form_data.get('telefone_2', [''])[0],
        "email": form_data.get('email', [''])[0],
        "capital_social": form_data.get('capital_social', [''])[0],
        "cnpj": form_data.get('cnpj', [''])[0],
        "socios": [
            {
                "nome": form_data.get('socios_nome[]', [''])[i],
                "qualificacao": form_data.get('socios_qualificacao[]', [''])[i],
            }
            for i in range(len(form_data.get('socios_nome[]', [])))
        ]
    }

    file_path = get_user_session_file()
    # Lê os dados do arquivo JSON e adiciona a nova empresa
    with open(file_path, 'r+') as file:
        empresas = json.load(file)
        empresas.append(empresa)
        file.seek(0)
        json.dump(empresas, file)

    return jsonify({'success': True, 'message': 'Empresa adicionada com sucesso.'})

def normalizar_cnpj(cnpj):
    return re.sub(r'\D', '', cnpj)

@app.route('/salvar_todas_csv', methods=['POST'])
def salvar_todas_csv():
    form_data = request.get_json()
    empresas = form_data.get('empresas', [])

    file_path = get_user_session_file()

    # Lê os CNPJs da blacklist
    blacklist_cnpjs = set()
    try:
        with open('resource/blacklist.csv', mode='r') as blacklist_file:
            csv_reader = csv.DictReader(blacklist_file)
            for row in csv_reader:
                cnpj_blacklist = normalizar_cnpj(row['cnpj'].strip())
                blacklist_cnpjs.add(cnpj_blacklist)
    except FileNotFoundError:
        return jsonify({'success': False, 'message': 'Arquivo de blacklist não encontrado.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

    # Debug: imprime CNPJs da blacklist
    print("CNPJs na blacklist:", blacklist_cnpjs)

    # Lê os dados existentes no arquivo JSON
    try:
        with open(file_path, 'r+') as file:
            try:
                empresas_cache = json.load(file)
            except json.JSONDecodeError:
                empresas_cache = []

            for empresa in empresas:
                cnpj = empresa.get('cnpj')
                cnpj_normalizado = normalizar_cnpj(cnpj)

                if cnpj_normalizado and cnpj_normalizado not in blacklist_cnpjs:
                    # Remove os acentos dos campos de cada empresa e sócios
                    empresa['razao_social'] = unidecode(empresa['razao_social'])
                    empresa['nome_fantasia'] = unidecode(empresa['nome_fantasia'])
                    empresa['logradouro'] = unidecode(empresa['logradouro'])

                    # Remover acentos dos sócios
                    for socio in empresa['socios']:
                        socio['nome'] = unidecode(socio['nome'])
                        socio['qualificacao'] = unidecode(socio['qualificacao'])

                    empresas_cache.append(empresa)
                else:
                    print("Ignorando CNPJ (Já processado):", cnpj_normalizado)

            file.seek(0)
            json.dump(empresas_cache, file, indent=4)
            file.truncate()

    except FileNotFoundError:
        return jsonify({'success': False, 'message': 'Arquivo de cache não encontrado.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

    return jsonify({'success': True, 'message': 'Todas as empresas, novas, foram salvas com sucesso.'})

def scrap_results_consolidated(resultados, scraps):
    consolidated_results = []

    for resultado, scrap in zip(resultados, scraps):
        # Dados do 'Resultado' (Google Maps)
        title_maps = resultado.get('title_maps')
        address_maps = resultado.get('address_maps')
        located_maps = resultado.get('located_maps')
        website = resultado.get('website_maps')
        phone1 = resultado.get('phone_maps')  # Phone do Google Maps
        page_url_maps = resultado.get('page_url_maps')

        # Dados do 'Scrap' (conhecimento gráfico)
        scrap_data = json.loads(scrap)
        knowledge_graph = scrap_data.get('knowledge_graph', {})
        consolidated_contact_info = scrap_data.get('consolidated_contact_info', {})

        # Rating e review_count
        rating = knowledge_graph.get('rating')
        review_count = knowledge_graph.get('review_count')

        # Endereço e segundo telefone (phone2)
        address = consolidated_contact_info.get('address')
        phone2 = consolidated_contact_info.get('phone')  # Phone do knowledge_graph ou consolidated_contact_info

        # Regras de prioridade para o endereço
        if address_maps == address:
            address = None
        elif not address_maps:
            address = address_maps

        # E-mail e perfis sociais
        email = consolidated_contact_info.get('email', None)
        social_media_profiles = consolidated_contact_info.get('social_media_profiles', [])

        # Criação do dicionário consolidado
        consolidated_info = {
            "title_maps": title_maps,
            "address_maps": address_maps or None,
            "located_maps": located_maps,
            "website": website,
            "phone1": phone1,   # Telefone do Google Maps
            "phone2": phone2,   # Telefone do knowledge_graph ou consolidated_contact_info
            "page_url_maps": page_url_maps,
            "rating": rating,
            "review_count": review_count,
            "address": address,
            "email": email if email else None,
            "social_media_profiles": social_media_profiles or None,
            "hours": None
        }

        consolidated_results.append(consolidated_info)

    return consolidated_results

@app.route('/baixar_csv')
def baixar_csv():
    file_path = get_user_session_file()

    # Lê os dados do arquivo temporário
    with open(file_path, 'r') as file:
        empresas = json.load(file)

    if not empresas:
        return jsonify({'error': 'Nenhuma empresa disponível para download.'}), 400

    # Faz o scrap dos dados de forma assíncrona
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    queries = []

    for empresa in empresas:
        # Usa o nome fantasia se disponível, caso contrário a razão social
        nome = empresa['nome_fantasia'] if empresa['nome_fantasia'] else empresa['razao_social']

        # Extrai o município do logradouro
        logradouro_parts = empresa['logradouro'].split(" - ")
        municipio = logradouro_parts[-2] if len(logradouro_parts) > 1 else ''  # Assume penúltimo item como município

        # Concatena nome e município
        query = f"{nome} {municipio}".strip()
        queries.append(query)
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    scrap_results = loop.run_until_complete(integration.scrap.process_queries(queries))

    scrap_results_maps = realizar_pesquisas(queries)
    resultados_consolidados = scrap_results_consolidated(scrap_results_maps, scrap_results)
    # Blacklist
    blacklist_path = 'resource/blacklist.csv'

    # Verifica se o arquivo já existe, senão cria com cabeçalhos
    if not os.path.exists(blacklist_path):
        with open(blacklist_path, 'w', newline='') as blacklist_file:
            writer = csv.writer(blacklist_file)
            writer.writerow(['cnpj', 'user', 'datetime'])  # Adiciona cabeçalhos

    # Nome do usuário logado e data/hora atual
    username = session.get('username', 'Usuário desconhecido')
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Adiciona os CNPJs no CSV de blacklist
    with open(blacklist_path, 'a', newline='') as blacklist_file:
        writer = csv.writer(blacklist_file)
        for empresa in empresas:
            cnpj_normalizado = normalizar_cnpj(empresa['cnpj'])  # Normaliza o CNPJ
            writer.writerow([cnpj_normalizado, username, current_time])

    # Chama a função para gerar o Excel
    output = gerar_excel(empresas, resultados_consolidados)

    # Remover o arquivo temporário após o download
    os.remove(file_path)
    session.pop('user_id', None)

    # Retorna o XLSX como resposta
    return Response(output,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=empresas.xlsx"})


@app.route('/enviar_empresa', methods=['POST'])
def enviar_empresa():
    empresa = {
        "razao_social": request.form.get('razao_social'),
        "nome_fantasia": request.form.get('nome_fantasia'),
        "logradouro": request.form.get('logradouro'),
        "municipio": request.form.get('municipio'),
        "uf": request.form.get('uf'),
        "cep": request.form.get('cep'),
        "telefone_1": request.form.get('telefone_1'),
        "telefone_2": request.form.get('telefone_2'),
        "email": request.form.get('email'),
        "porte": request.form.get('porte'),
        "cnpj": request.form.get('cnpj'),
        "socios": [
            {
                "nome": nome_socio,
                "faixa_etaria": faixa_etaria,
                "qualificacao": qualificacao,
                "data_entrada": data_entrada
            }
            for nome_socio, faixa_etaria, qualificacao, data_entrada in zip(
                request.form.getlist('socios_nome[]'),
                request.form.getlist('socios_faixa_etaria[]'),
                request.form.getlist('socios_qualificacao[]'),
                request.form.getlist('socios_data_entrada[]')
            )
        ]
    }

    usuario_logado = session.get('username', 'Usuário Desconhecido')

    lead_existente = integration.verificar_lead_existente_por_titulo(f"Via Automação - {empresa['razao_social']} - CNPJ: {empresa['cnpj']}")

    if lead_existente:
        lead_id = lead_existente[0]['ID']
        lead_link = f"https://setup.bitrix24.com.br/crm/lead/show/{lead_id}/"
        return jsonify({"success": True, "lead_existente": True, "lead_link": lead_link})
    else:
        # Passar o nome do usuário para a função do integration.py
        integration.enviar_dados_bitrix([empresa], usuario_logado)
        return jsonify({"success": True, "lead_existente": False})

@app.route('/logout', methods=['POST'])
def logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    session.pop('full_name', None)
    session.pop('phone_number', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=5000)
